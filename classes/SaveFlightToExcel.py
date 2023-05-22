from openpyxl import Workbook, load_workbook
from classes.Flight import Flight
import os
import logging

class SaveFlightToExcel:
    title = "Flights"

    def __init__(self,  filename):
        self.filename = filename

        if os.path.exists(self.filename):
            try:
                self.wb = load_workbook(self.filename)
            except Exception as e:
                logging.error(e)
                print("--ERROR-- loading excel file")
        else:
            try:
                self.wb = Workbook()
                ws = self.wb.active
                ws.title = self.title
                ws.append(["Departure", "Arrival", "Departure Date", "Arrival Date", "Travel Time", "Price", "Link"])
            except Exception as e:
                logging.error(e)
                print("--ERROR-- creating excel file")

    def save(self, flight: Flight):
        try:
            ws = self.wb[self.title]
            ws.append([flight.departure, flight.arrival, flight.departure_date, flight.arrival_date, flight.travel_time, flight.price, flight.link])
            self.wb.save(self.filename)
        except Exception as e:
            logging.error(e)
            print("--ERROR-- saving flight to excel")

